import requests
import json
from openpyxl import Workbook
from prettytable import PrettyTable
from datetime import datetime, timedelta
import pandas as pd


class GetTrains:
    def __init__(self, date, begin_id, end_id):
        # 请求的目标链接
        self.url = "https://kyfw.12306.cn/otn/leftTicket/query"
        # cookies
        self.cookies = {
            '_uab_collina': '171324859263120074949415',
            'JSESSIONID': '708D9C6917F9858184F462E86DC45BD0',
            '_jc_save_fromStation': '%u82CF%u5DDE%2CSZH',
            '_jc_save_toStation': '%u6C5D%u5DDE%2CROF',
            '_jc_save_fromDate': '2024-04-30',
            '_jc_save_wfdc_flag': 'dc',
            'route': '9036359bb8a8a461c164a04f8f50b252',
            'BIGipServerotn': '1172832522.24610.0000',
            'BIGipServerpassport': '854065418.50215.0000',
            'guidesStatus': 'off',
            'highContrastMode': 'defaltMode',
            'cursorStatus': 'off',
            '_jc_save_toDate': '2024-04-30',
        }
        # 构建请求头
        self.headers = {
            'Accept': '*/*',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6,zh-TW;q=0.5',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'If-Modified-Since': '0',
            'Pragma': 'no-cache',
            'Referer': 'https://www.12306.cn/index/index.html',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
            'X-Requested-With': 'XMLHttpRequest',
        }
        # 构建请求所需参数
        self.params = {
            "leftTicketDTO.train_date": date,
            "leftTicketDTO.from_station": begin_id,
            "leftTicketDTO.to_station": end_id,
            "purpose_codes": "ADULT"
        }
        # 实例化美化表格对象
        self.pt = PrettyTable()

    def run(self):
        # 对目标网址发送请求
        res = requests.get(
            self.url, headers=self.headers, params=self.params, cookies=self.cookies
        ).json()
        data_list = res['data']['result']
        # 构造表格的表头，用于展示和保存
        header_list = [
            ['车次', '出发时间', '到达时间', '历时', '商务座剩余票数', '一等座剩余票数', '二等座剩余票数', '软卧剩余票数', '硬卧剩余票数', '硬座剩余票数', '无座剩余票数',
             '备注']
        ]
        # 将表头信息添加进展示表格的表头
        self.pt.field_names = header_list[0]
        for data in data_list:
            # 格式化添加表数据
            trains_msg = self.format_data(data)
            header_list.append(trains_msg)
        # 打印表格
        # print(self.pt)
        # 返回车次信息列表
        return header_list

    def format_data(self, data):
        # 将返回的数据以'|'进行分隔
        all_data_list = data.split('|')
        # 提取车次的信息
        trains_msg = [
            all_data_list[3],
            all_data_list[8],
            all_data_list[9],
            all_data_list[10],
            all_data_list[32] if all_data_list[32] != "" else "--",
            all_data_list[31] if all_data_list[31] != "" else "--",
            all_data_list[30] if all_data_list[30] != "" else "--",
            all_data_list[23] if all_data_list[23] != "" else "--",
            all_data_list[28] if all_data_list[28] != "" else "--",
            all_data_list[29] if all_data_list[29] != "" else "--",
            all_data_list[26] if all_data_list[26] != "" else "--",
            all_data_list[1] if all_data_list[1] != "" else "--"
        ]
        # 增添表内容
        self.pt.add_row(trains_msg)
        # 将提取的信息返回，用于保存
        return trains_msg

    def save_data(self, trains_data_list, date, begin, end):
        num = input("如果展示不清晰，需要保存时请扣1：")
        if num == "1":
            wb = Workbook()
            sheet = wb.create_sheet("车次信息", -1)
            # 遍历表格索引，写入数据
            for x in range(len(trains_data_list)):
                for y in range(len(trains_data_list[x])):
                    sheet.cell(x + 1, y + 1).value = trains_data_list[x][y]
            wb.save(f"{date}_{begin}_{end}.xlsx")
            print("数据保存完成！")


# 返回给你的是一个list，内容分别是
#             ['车次', '出发时间', '到达时间', '历时', '商务座', '一等座', '二等座', '软卧', '硬卧', '硬座', '无座', '备注']
def get_12306(date, begin, end):
    # 检查日期是否有效
    input_date = datetime.strptime(date, "%Y-%m-%d").date()
    today = datetime.today().date()
    half_month_later = today + timedelta(days=14)
    print(today, half_month_later)
    if not today <= input_date <= half_month_later:
        print("日期无效，必须是今天或以后半个月之内的日期")
        return pd.DataFrame()
    # 读取城市信息json文件
    city_list = json.load(open('./data/city_data.json', 'r', encoding='GBK'))
    # 获取城市对应的英文代码
    begin_id = city_list[begin]
    end_id = city_list[end]
    gt = GetTrains(date, begin_id, end_id)
    trains_data_list = gt.run()
    if len(trains_data_list) == 1:
        print("没有查询到车次信息！")
        return pd.DataFrame()
    return pd.DataFrame(trains_data_list[1:], columns=trains_data_list[0])

def mmy_12306(date, begin, end):
    a = get_12306(date, begin, end)
    if a.empty:
        return "没有查询到车次信息！"
    else:
        a.drop(columns=["备注"], inplace=True)
        a.sort_values(by="历时", inplace=True)
        formatted_list = a.apply(lambda row: ', '.join(['{}: {}'.format(col, val) for col, val in row.to_dict().items() if val not in ["无", "--"]]) if list(row).count("无") + list(row).count("--") < 7 else '', axis=1).tolist()
        formatted_list = [i for i in formatted_list if i]  # remove empty strings
        return formatted_list[0:5]
    # a.to_excel("trains_data.xlsx", index=False, engine='openpyxl')

if __name__ == '__main__':
    a = mmy_12306("2024-06-30", "天津", "哈尔滨")
    print(a)
    # pass

